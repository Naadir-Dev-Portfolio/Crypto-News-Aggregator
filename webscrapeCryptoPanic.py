from PyQt6.QtWidgets import QApplication, QMessageBox
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
import sys

# Initialize PyQt application
app = QApplication(sys.argv)

def show_message_box():
    msg = QMessageBox()
    msg.setWindowTitle("Ready to Scrape?")
    msg.setText("Please scroll through the webpage to ensure all articles are loaded. Click OK when ready to scrape.")
    msg.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
    return msg.exec()

# Initialize WebDriver
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

driver.get('https://cryptopanic.com/')
print("Please switch to the browser and scroll to load all articles.")
user_response = show_message_box()

if user_response == QMessageBox.StandardButton.Ok:
    def collect_articles(driver):
        articles_data = []
        news_articles = driver.find_elements(By.CSS_SELECTOR, ".news-row.news-row-link")
        for article in news_articles:
            date_time_str = article.find_element(By.CSS_SELECTOR, "time").get_attribute('datetime')
            date_time_obj = datetime.strptime(date_time_str, "%a %b %d %Y %H:%M:%S GMT+0000 (Greenwich Mean Time)")
            link = article.find_element(By.CSS_SELECTOR, "a.news-cell.nc-title").get_attribute('href')
            articles_data.append([date_time_obj.strftime('%Y-%m-%d %H:%M:%S'), link])
        
        return articles_data

    articles = collect_articles(driver)
    driver.quit()

    # Create an XLSX file and add data
    wb = Workbook()
    ws = wb.active
    ws.append(['Date and Time', 'Link', 'Generated Headline'])

    for row in articles:
        ws.append(row)
        last_row = ws.max_row
        link_cell = 'B' + str(last_row)
        formula = f"=SUBSTITUTE(SUBSTITUTE(MID({link_cell}, FIND(\"☺\", SUBSTITUTE({link_cell}, \"/\", \"☺\", LEN({link_cell}) - LEN(SUBSTITUTE({link_cell}, \"/\", \"\")))), LEN({link_cell})), \"-\", \" \"), \"/\", \"\")"
        ws[f'C{last_row}'] = formula

    xlsx_file_path = 'cryptopanic_articles.xlsx'
    wb.save(xlsx_file_path)

    print(f"XLSX file has been created: {xlsx_file_path}")
else:
    print("Scraping cancelled.")
    driver.quit()
